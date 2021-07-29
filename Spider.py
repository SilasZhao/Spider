import requests
import json
import time
import openpyxl

url = "https://report.amap.com/ajax/cityHourly.do?cityCode={}&dataType=4"

city_name_list = list(["北京市","上海市","广州市","深圳市","成都市","杭州市","武汉市"])

#从excel中找到citycode
def getCityCode(city,filename):
    try:
        file = openpyxl.load_workbook(filename)
        f = file.active
    except Exception as e:
        print(e)
        exit(1)
    else:
        for i in range(1,f.max_row+1): 
            if f.cell(row=i,column=1).value == city:
                return f.cell(row=i,column=2).value

#利用city_name_list生成cityList
def getCityList(city_name_list,filename):
    cityList = []
    for i in city_name_list:
        l = list([i,getCityCode(i,filename)])
        cityList.append(l)
    return cityList

#尝试利用cityCode取得该城市的平均时速的列表
def getList(url,cityCode):
    try:
        response = requests.get(url.format(cityCode))
        response.raise_for_status()
    except requests.RequestException as e:
        print(e)
        exit(1)
    else:
        return json.loads(response.text)

#从平均时速列表中返回目标时间的平均时速
#在8-9点时没有数据
def getNumber(targetTime,responseList):
    for i in responseList:
        #每个list的list[0]/1000就是其时间的时间戳
        timeStamp = int((i[0] / 1000))
        hour = time.localtime(timeStamp).tm_hour
        if hour == targetTime:
            #返回车速
            return str(i[1])

if __name__=='__main__':
    code = []
    cityList = getCityList(city_name_list,"cityCode.xlsx")
    for i in cityList:
        speed = {i[0]: getNumber(8,getList(url,i[1]))}
        code.append(speed)
    print(json.dumps(code,ensure_ascii=False))



    

